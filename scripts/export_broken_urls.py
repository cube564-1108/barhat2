"""
Excel-выгрузка товаров, у которых ссылка в RetailCRM расходится с сайтом.

Для работы в админке Битрикса: артикул, наименование, что именно не так,
какая ссылка сейчас и какая должна быть. Правильную ссылку берём не из догадок,
а из <link rel="canonical"> самой страницы товара.

Технический раздел «Товары МС» (`tovary-ms`) исключён: этих позиций на сайте нет
в принципе, чинить там нечего — их надо убирать из выгрузки целиком (шаг 1
инструкции docs/инструкция-ссылки-товаров-битрикс-crm.md).

Обход последовательный: карточка генерится ~4 секунды, при параллельной проверке
сайт отдаёт 503 (см. комментарий в scripts/check_product_urls.py). Полный проход
по ~350 товарам занимает около получаса.

Запуск:
    python scripts/export_broken_urls.py
    python scripts/export_broken_urls.py --limit 30   # проверить на выборке
"""

import argparse
import os
import re
import sys
import time

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'src'))

from dotenv import load_dotenv

REPO_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
load_dotenv(os.path.join(REPO_ROOT, '.env'))

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from russian_ca import trust_russian_ca

RETAILCRM_URL = (os.environ.get('RETAILCRM_URL') or '').rstrip('/')
RETAILCRM_API_KEY = os.environ.get('RETAILCRM_API_KEY') or ''

CRM_TIMEOUT = 30
SITE_TIMEOUT = 25
PAGE_LIMIT = 100
PAUSE = 0.1

# Раздел-исключение: номенклатура из МойСклада, на сайте её нет
EXCLUDED_SECTION = '/tovary-ms/'

OUTPUT_PATH = os.path.join(REPO_ROOT, 'data', 'ссылки-товаров-расхождения.xlsx')

CANONICAL_RE = re.compile(
    r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)["\']', re.I
)


def fetch_products(limit=None):
    if not RETAILCRM_URL or not RETAILCRM_API_KEY:
        sys.exit('RETAILCRM_URL или RETAILCRM_API_KEY не заданы в .env')

    crm = requests.Session()
    crm.headers.update({'X-API-KEY': RETAILCRM_API_KEY})
    trust_russian_ca(crm)

    products, page = [], 1
    while True:
        response = crm.get(
            f'{RETAILCRM_URL}/api/v5/store/products',
            params={'limit': PAGE_LIMIT, 'page': page},
            timeout=CRM_TIMEOUT,
        )
        response.raise_for_status()
        data = response.json()
        products += data.get('products', [])

        if limit and len(products) >= limit:
            return products[:limit]
        if page >= data.get('pagination', {}).get('totalPageCount', 1):
            return products
        page += 1
        time.sleep(0.2)


def _leaf(url):
    return url.replace(':443', '').rstrip('/').split('/')[-1]


def inspect(product, session):
    """Одна ссылка → (что не так, правильная ссылка) либо None, если всё хорошо."""
    url = product.get('url') or ''

    try:
        response = session.get(url, timeout=SITE_TIMEOUT, allow_redirects=True)
    except requests.exceptions.RequestException as e:
        return f'Ошибка запроса: {type(e).__name__}', ''

    final = response.url.replace(':443', '')

    if response.status_code == 404:
        return 'Страница не найдена (404)', ''
    if not response.ok:
        return f'Сайт ответил {response.status_code}', ''

    # Товар не найден — Битрикс показал раздел вместо карточки. Правильной
    # ссылки у такого товара нет вообще, поэтому колонка остаётся пустой:
    # подставить туда адрес раздела значило бы предложить отправлять клиенту его.
    if _leaf(url) != _leaf(final):
        return 'Товара нет на сайте — открывается раздел', ''

    canonical = CANONICAL_RE.search(response.text)
    right = canonical.group(1) if canonical else final

    if right.rstrip('/') != url.rstrip('/'):
        return 'Ссылка неканоническая — открывается через редирект', right

    return None, right


def write_excel(rows, checked):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Расхождения'

    headers = [
        'Артикул', 'Наименование', 'Что не так',
        'Ссылка сейчас (в CRM)', 'Правильная ссылка', 'ID в CRM',
    ]
    header_fill = PatternFill('solid', fgColor='7A1F3D')
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for row in rows:
        ws.append(row)

    widths = [12, 46, 44, 62, 62, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--limit', type=int, help='проверить только первые N товаров')
    args = parser.parse_args()

    print('Забираю каталог из RetailCRM...')
    products = fetch_products(args.limit)

    excluded = [p for p in products if EXCLUDED_SECTION in (p.get('url') or '')]
    products = [p for p in products if EXCLUDED_SECTION not in (p.get('url') or '')]
    print(f'товаров: {len(products)} (исключено «Товары МС»: {len(excluded)})')
    print('Проверяю ссылки, это небыстро...')

    site = requests.Session()
    site.headers.update({'User-Agent': 'barhat-link-check/1.0'})
    trust_russian_ca(site)

    rows = []
    for done, product in enumerate(products, start=1):
        problem, right = inspect(product, site)
        if problem:
            rows.append([
                product.get('article') or '',
                product.get('name') or '',
                problem,
                product.get('url') or '',
                right,
                product.get('id'),
            ])
        if done % 25 == 0:
            print(f'  проверено {done} из {len(products)}, расхождений {len(rows)}')
        time.sleep(PAUSE)

    rows.sort(key=lambda r: (r[2], r[1]))
    write_excel(rows, len(products))

    print(f'\nПроверено товаров: {len(products)}')
    print(f'С расхождением:    {len(rows)}')
    print(f'\nФайл: {OUTPUT_PATH}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
